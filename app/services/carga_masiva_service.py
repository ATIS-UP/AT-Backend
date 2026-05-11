"""service layer for bulk student upload from csv/xlsx files"""
import csv
import io
from typing import Optional

from fastapi import UploadFile
from sqlalchemy.orm import Session

from app.exceptions import ValidationError
from app.models.estudiante import Estudiante, EstadoEstudiante
from app.utils.security import encrypt_data
from app.utils.audit import AuditService

# required columns that must be present in the uploaded file
REQUIRED_COLUMNS = {"codigo", "nombres", "apellidos", "programa", "semestre"}

# valid file extensions
ALLOWED_EXTENSIONS = {".csv", ".xlsx"}

# valid student states
VALID_ESTADOS = {"ACTIVO", "INACTIVO", "GRADUADO", "SUSPENDIDO"}


class CargaMasivaService:
    """handles bulk upload of students from csv/xlsx files"""

    def __init__(self, db: Session):
        self.db = db

    def procesar_archivo(
        self, file: UploadFile, usuario_id: str, ip: str
    ) -> dict:
        """process an uploaded csv/xlsx file and insert/update students.

        returns a summary dict with total_filas, insertadas, actualizadas,
        errores, and detalle_errores.
        """
        extension = self._get_extension(file.filename)
        self._validate_extension(extension)

        rows = self._read_file(file, extension)
        self._validate_columns(rows)

        insertadas = 0
        actualizadas = 0
        errores = 0
        detalle_errores = []

        for idx, row in enumerate(rows, start=2):
            # row index starts at 2 (header is row 1)
            row_errors = self._validate_row(row, idx)
            if row_errors:
                errores += len(row_errors)
                detalle_errores.extend(row_errors)
                continue

            try:
                was_updated = self._upsert_student(row, usuario_id, ip)
                if was_updated:
                    actualizadas += 1
                else:
                    insertadas += 1
            except Exception as e:
                errores += 1
                detalle_errores.append(
                    {"fila": idx, "campo": "general", "error": str(e)}
                )

        # commit all changes at once
        self.db.commit()

        total_filas = len(rows)

        AuditService.log(
            db=self.db,
            usuario_id=usuario_id,
            accion="CARGA_MASIVA",
            entidad="Estudiante",
            detalles={
                "total_filas": total_filas,
                "insertadas": insertadas,
                "actualizadas": actualizadas,
                "errores": errores,
            },
            ip=ip,
            estado="EXITOSO",
            mensaje=f"Carga masiva: {insertadas} insertadas, {actualizadas} actualizadas, {errores} errores",
        )

        return {
            "total_filas": total_filas,
            "insertadas": insertadas,
            "actualizadas": actualizadas,
            "errores": errores,
            "detalle_errores": detalle_errores,
        }

    def _get_extension(self, filename: Optional[str]) -> str:
        """extract lowercase file extension from filename"""
        if not filename:
            return ""
        dot_idx = filename.rfind(".")
        if dot_idx == -1:
            return ""
        return filename[dot_idx:].lower()

    def _validate_extension(self, extension: str) -> None:
        """raise ValidationError if extension is not csv or xlsx"""
        if extension not in ALLOWED_EXTENSIONS:
            raise ValidationError(
                message="Formato de archivo no soportado. Use CSV o XLSX.",
                fields={"archivo": f"Extensión '{extension}' no permitida"},
            )

    def _read_file(self, file: UploadFile, extension: str) -> list[dict]:
        """read file content into a list of row dicts"""
        content = file.file.read()

        if extension == ".csv":
            return self._read_csv(content)
        elif extension == ".xlsx":
            return self._read_xlsx(content)
        return []

    def _read_csv(self, content: bytes) -> list[dict]:
        """parse csv content into list of dicts with normalized column names"""
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("latin-1")

        reader = csv.DictReader(io.StringIO(text))
        rows = []
        for row in reader:
            # normalize column names to lowercase and strip whitespace
            normalized = {
                k.strip().lower(): v.strip() if v else ""
                for k, v in row.items()
                if k is not None
            }
            rows.append(normalized)
        return rows

    def _read_xlsx(self, content: bytes) -> list[dict]:
        """parse xlsx content into list of dicts with normalized column names"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ValidationError(
                message="Soporte XLSX no disponible. Instale openpyxl.",
                fields={"archivo": "openpyxl no instalado"},
            )

        wb = load_workbook(filename=io.BytesIO(content), read_only=True)
        ws = wb.active

        rows_iter = ws.iter_rows(values_only=True)
        # first row is the header
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return []

        headers = [
            str(h).strip().lower() if h else "" for h in header_row
        ]

        rows = []
        for row_values in rows_iter:
            row_dict = {}
            for col_idx, value in enumerate(row_values):
                if col_idx < len(headers) and headers[col_idx]:
                    row_dict[headers[col_idx]] = (
                        str(value).strip() if value is not None else ""
                    )
            rows.append(row_dict)

        wb.close()
        return rows

    def _validate_columns(self, rows: list[dict]) -> None:
        """check that required columns exist in the parsed data"""
        if not rows:
            raise ValidationError(
                message="El archivo está vacío o no contiene datos.",
                fields={"archivo": "sin datos"},
            )

        available_columns = set(rows[0].keys())
        missing = REQUIRED_COLUMNS - available_columns
        if missing:
            raise ValidationError(
                message=f"Columnas requeridas faltantes: {', '.join(sorted(missing))}",
                fields={"columnas": list(sorted(missing))},
            )

    def _validate_row(self, row: dict, row_number: int) -> list[dict]:
        """validate a single row, return list of error dicts (empty if valid)"""
        errors = []

        # codigo: required, max 20 chars
        codigo = row.get("codigo", "").strip()
        if not codigo:
            errors.append(
                {"fila": row_number, "campo": "codigo", "error": "campo requerido"}
            )
        elif len(codigo) > 20:
            errors.append(
                {"fila": row_number, "campo": "codigo", "error": "máximo 20 caracteres"}
            )

        # nombres: required
        nombres = row.get("nombres", "").strip()
        if not nombres:
            errors.append(
                {"fila": row_number, "campo": "nombres", "error": "campo requerido"}
            )

        # apellidos: required
        apellidos = row.get("apellidos", "").strip()
        if not apellidos:
            errors.append(
                {"fila": row_number, "campo": "apellidos", "error": "campo requerido"}
            )

        # programa: required
        programa = row.get("programa", "").strip()
        if not programa:
            errors.append(
                {"fila": row_number, "campo": "programa", "error": "campo requerido"}
            )

        # semestre: required, integer between 1 and 15
        semestre_str = row.get("semestre", "").strip()
        if not semestre_str:
            errors.append(
                {"fila": row_number, "campo": "semestre", "error": "campo requerido"}
            )
        else:
            try:
                semestre_val = int(semestre_str)
                if semestre_val < 1 or semestre_val > 15:
                    errors.append(
                        {
                            "fila": row_number,
                            "campo": "semestre",
                            "error": "debe ser un entero entre 1 y 15",
                        }
                    )
            except ValueError:
                errors.append(
                    {
                        "fila": row_number,
                        "campo": "semestre",
                        "error": "debe ser un entero válido",
                    }
                )

        # estado: optional, must be one of valid states
        estado = row.get("estado", "").strip().upper()
        if estado and estado not in VALID_ESTADOS:
            errors.append(
                {
                    "fila": row_number,
                    "campo": "estado",
                    "error": f"debe ser uno de: {', '.join(sorted(VALID_ESTADOS))}",
                }
            )

        return errors

    def _upsert_student(
        self, row: dict, usuario_id: str, ip: str
    ) -> bool:
        """insert or update a student from row data.

        returns True if updated, False if inserted.
        """
        codigo = row["codigo"].strip()
        existing = (
            self.db.query(Estudiante)
            .filter(Estudiante.codigo == codigo)
            .first()
        )

        nombres = row.get("nombres", "").strip()
        apellidos = row.get("apellidos", "").strip()
        programa = row.get("programa", "").strip()
        semestre = int(row.get("semestre", "1").strip())
        estado_str = row.get("estado", "").strip().upper() or "ACTIVO"
        email = row.get("email", "").strip() or None
        documento = row.get("documento", "").strip() or None
        telefono = row.get("telefono", "").strip() or None

        estado = EstadoEstudiante(estado_str)

        if existing:
            # update existing student
            existing.nombres = encrypt_data(nombres)
            existing.apellidos = encrypt_data(apellidos)
            existing.programa = programa
            existing.semestre = semestre
            existing.estado = estado
            if email:
                existing.email = encrypt_data(email)
            if documento:
                existing.documento = encrypt_data(documento)
            if telefono:
                existing.telefono = encrypt_data(telefono)
            return True
        else:
            # insert new student
            nuevo = Estudiante(
                codigo=codigo,
                nombres=encrypt_data(nombres),
                apellidos=encrypt_data(apellidos),
                programa=programa,
                semestre=semestre,
                estado=estado,
                email=encrypt_data(email) if email else None,
                documento=encrypt_data(documento) if documento else None,
                telefono=encrypt_data(telefono) if telefono else None,
            )
            self.db.add(nuevo)
            return False
